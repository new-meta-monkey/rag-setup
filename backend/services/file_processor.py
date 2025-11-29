import io
from typing import Optional
from fastapi import UploadFile, HTTPException
import PyPDF2
import docx
import openpyxl
from backend.utils.text_cleaner import TextCleaner

class FileProcessor:
    @staticmethod
    async def extract_text(file: UploadFile, use_unstructured: bool = True) -> dict:
        """
        Extract text from uploaded file based on content type or extension
        
        Args:
            file: Uploaded file
            use_unstructured: Whether to use unstructured library for better extraction
                             (handles HTML removal, structure preservation automatically)
        
        Returns:
            dict: {
                "text": "full cleaned text",
                "pages": [{"page_number": 1, "text": "..."}]
            }
        """
        content = await file.read()
        filename = file.filename.lower()
        
        try:
            # Option 1: Use unstructured library for advanced partitioning and cleaning
            if use_unstructured and (filename.endswith('.pdf') or filename.endswith('.docx') or 
                                     filename.endswith('.html') or filename.endswith('.htm')):
                try:
                    return TextCleaner.partition_and_clean_with_pages(content, file.filename)
                except Exception as e:
                    # Fallback to legacy extraction if unstructured fails
                    print(f"Unstructured partitioning failed: {e}. Using legacy extraction.")
            
            # Option 2: Legacy extraction with text cleaning
            pages = []
            raw_text = ""
            
            if filename.endswith('.pdf'):
                pages = FileProcessor._extract_pdf(content)
                raw_text = "\n\n".join([p["text"] for p in pages])
            elif filename.endswith('.docx'):
                raw_text = FileProcessor._extract_docx(content)
                pages = [{"page_number": 1, "text": raw_text}]
            elif filename.endswith('.xlsx'):
                raw_text = FileProcessor._extract_xlsx(content)
                pages = [{"page_number": 1, "text": raw_text}]
            elif filename.endswith('.txt') or filename.endswith('.md'):
                raw_text = content.decode('utf-8')
                pages = [{"page_number": 1, "text": raw_text}]
            elif filename.endswith('.html') or filename.endswith('.htm'):
                raw_text = content.decode('utf-8', errors='ignore')
                # Remove HTML tags for HTML files
                raw_text = TextCleaner.remove_html_tags(raw_text)
                pages = [{"page_number": 1, "text": raw_text}]
            else:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Unsupported file type: {filename}. Supported types: .pdf, .docx, .xlsx, .txt, .md, .html"
                )
            
            # Clean the extracted text
            cleaned_text = TextCleaner.clean_text(raw_text)
            
            # Clean pages text as well
            cleaned_pages = []
            for page in pages:
                cleaned_pages.append({
                    "page_number": page["page_number"],
                    "text": TextCleaner.clean_text(page["text"])
                })
                
            return {
                "text": cleaned_text,
                "pages": cleaned_pages
            }
            
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

    @staticmethod
    def _extract_pdf(content: bytes) -> list:
        pages = []
        pdf_file = io.BytesIO(content)
        reader = PyPDF2.PdfReader(pdf_file)
        
        for i, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            pages.append({
                "page_number": i + 1,
                "text": text
            })
            
        return pages

    @staticmethod
    def _extract_docx(content: bytes) -> str:
        doc_file = io.BytesIO(content)
        doc = docx.Document(doc_file)
        return "\n\n".join([paragraph.text for paragraph in doc.paragraphs])

    @staticmethod
    def _extract_xlsx(content: bytes) -> str:
        text = []
        xlsx_file = io.BytesIO(content)
        workbook = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text.append(f"--- Sheet: {sheet_name} ---")
            for row in sheet.iter_rows(values_only=True):
                # Filter out None values and convert to string
                row_text = [str(cell) for cell in row if cell is not None]
                if row_text:
                    text.append(" | ".join(row_text))
            text.append("\n")
            
        return "\n".join(text)
